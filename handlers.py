from typing import Generator, Dict, Iterable, List
import time

from openpyxl.styles.cell_style import StyleArray
from openpyxl.worksheet.worksheet import Worksheet, Cell


class Header:
    def __init__(self, row: Iterable = None):
        self._names = {}
        self._idxes = {}
        self._row = row if row else []
        self.row = self._row

    @property
    def row(self) -> Iterable:
        return self._row

    @row.setter
    def row(self, other) -> None:
        for cell in other:
            val = cell.value
            idx = cell.col_idx
            self._names[val] = idx
            self._idxes[idx] = val

    def get_name_by_idx(self, idx: int) -> str:
        return self._idxes.get(idx, None)

    def get_idx_by_name(self, name: str) -> int:
        return self._names.get(name, None)


class WorkSheetHandler:
    CODE_TITLE = "#2"

    def __init__(self, _workbook, sheet_name):
        self._sheet_name = ''
        self.sheet_name: str = sheet_name
        self.workbook = _workbook
        self.sheet: Worksheet = self.workbook[self.sheet_name]
        self.header = Header(next(self.sheet.iter_rows(0, 1)))
        self.code_data = {}
        self.code_column = []

    @property
    def sheet_name(self) -> str:
        return self._sheet_name

    @property
    def rows(self) -> Generator:
        return self.sheet.iter_rows(min_row=2)

    @sheet_name.setter
    def sheet_name(self, name: str) -> None:
        if not isinstance(name, str):
            raise ValueError("시트이름이 문자열타입이 아닙니다")
        self._sheet_name = name

    def get_title(self, target_cell: Cell) -> str:
        idx: int = target_cell.col_idx
        target_cell: Cell = self.sheet.cell(row=1, column=idx)
        return target_cell.value

    @property
    def code_idx(self) -> int:
        return self.header.get_idx_by_name(self.CODE_TITLE) - 1

    def get_code_column(self) -> List[Cell]:
        if not self.code_column:
            self.code_column = list(list(filter(lambda x: x[0].value == self.CODE_TITLE, self.sheet.iter_cols()))[0])
        return self.code_column

    def get_row_idx_by_code(self, code):
        if not self.code_data:
            column = self.get_code_column()
            self.code_data = {cell.value: cell.row for cell in column}
        return self.code_data[code]

    def copy_styles(self, new_handler: 'WorkSheetHandler'):
        for row in self.rows:
            row_hdr = RowHandler(row, self.code_idx)
            for cell in row:
                title = self.get_title(cell)
                style_hdr = StyleHandler(cell)
                if style_hdr.has_style and row_hdr.code:
                    row_hdr.names[title] = style_hdr.style

            row_idx: int = new_handler.get_row_idx_by_code(code=row_hdr.code)
            if not row_hdr.names:
                a = 1
            for k, v in row_hdr.names.items():
                col_idx: int = new_handler.header.get_idx_by_name(k)
                new_cell: Cell = new_handler.sheet.cell(row_idx, col_idx)
                new_cell._style = v
                new_cell._style.fillId = new_cell._style.fillId + 1
        return new_handler


class StyleHandler:
    def __init__(self, cell):
        self.style = cell._style

    @property
    def has_style(self):
        WHITE_ID = 4
        return self.style and self.style.fillId != WHITE_ID


class RowHandler:
    def __init__(self, row, code_idx):
        self._row: List[Cell] = row
        self._code_idx = code_idx
        self.names: Dict[str, StyleArray] = {}

    @property
    def code(self):
        return self._row[self._code_idx].value


if __name__ == '__main__':

    org_filename = '210215.xlsx'
    new_filename = 'temp.xlsx'

